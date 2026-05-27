# admin/export.py
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from typing import List, Dict
from utils.logger import logger

def create_excel_report(data: List[Dict], user_id: int, period_days: int = 30) -> io.BytesIO:
    """
    Генерирует Excel-отчёт с результатами викторин.
    Возвращает BytesIO объект для отправки в Telegram.
    """
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Отчёт_{user_id}"
    
    # 🔹 Заголовки с форматированием
    headers = [
        "📅 Дата", "📁 Файл", "❓ Вопрос", "✅ Правильный ответ",
        "👤 Ответ пользователя", "🎯 Результат", "⭐ Оценка", 
        "💡 Пояснение ИИ", "🪙 Токены (ген+оценка)"
    ]
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25
    
    # 🔹 Данные
    for row_idx, record in enumerate(data, start=2):
        ws.cell(row=row_idx, column=1, value=record['generated_at'][:10])  # Дата
        ws.cell(row=row_idx, column=2, value=record['source_file'])         # Файл
        ws.cell(row=row_idx, column=3, value=record['question'])            # Вопрос
        ws.cell(row=row_idx, column=4, value=record['correct_answer'])      # Правильный ответ
        ws.cell(row=row_idx, column=5, value=record['user_answer'] or "—")  # Ответ пользователя
        
        # 🎨 Цветовая индикация результата
        correctness = record['correctness'] or "—"
        cell = ws.cell(row=row_idx, column=6, value=correctness)
        if correctness == "правильно":
            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif correctness == "частично":
            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        elif correctness == "неправильно":
            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        ws.cell(row=row_idx, column=7, value=record['rating'] or "—")       # Оценка
        ws.cell(row=row_idx, column=8, value=record['feedback'] or "—")     # Пояснение ИИ
        tokens = f"{record['gen_tokens'] + record['eval_tokens']}"
        ws.cell(row=row_idx, column=9, value=tokens)                        # Токены
        
        # 🔹 Перенос текста для длинных ячеек
        for col in range(3, 9):  # Вопрос, ответы, фидбек
            ws.cell(row=row_idx, column=col).alignment = Alignment(wrap_text=True)
    
    # 🔹 Итоговая статистика на новом листе
    stats_ws = wb.create_sheet("📈 Статистика")
    total = len(data)
    correct = sum(1 for r in data if r['correctness'] == 'правильно')
    avg_rating = sum(r['rating'] for r in data if r['rating']) / max(1, sum(1 for r in data if r['rating']))
    total_tokens = sum((r['gen_tokens'] or 0) + (r['eval_tokens'] or 0) for r in data)
    
    stats = [
        ["📊 Период", f"Последние {period_days} дней"],
        ["🔢 Всего вопросов", total],
        ["✅ Правильных ответов", f"{correct} ({correct/total*100:.1f}%)"],
        ["⭐ Средняя оценка", f"{avg_rating:.2f}"],
        ["🪙 Всего токенов", total_tokens],
        ["👤 User ID", user_id],
        ["🕐 Дата выгрузки", datetime.now().strftime("%Y-%m-%d %H:%M")]
    ]
    
    for row_idx, (label, value) in enumerate(stats, start=1):
        stats_ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        stats_ws.cell(row=row_idx, column=2, value=value)
        stats_ws.column_dimensions['A'].width = 25
        stats_ws.column_dimensions['B'].width = 30
    
    # 🔹 Сохранение в буфер
    wb.save(output)
    output.seek(0)
    logger.info(f"📄 Excel report generated for user {user_id}: {total} questions")
    return output

